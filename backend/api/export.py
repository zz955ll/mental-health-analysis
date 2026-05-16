"""
报告导出模块 - PDF/Excel 导出（纯英文版）
"""

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm


def export_to_excel(data, filename='prediction_report.xlsx'):
    """导出到Excel（支持中文）"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Prediction Result'

    ws1['A1'] = 'Depression Risk Prediction Report'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:D1')

    ws1['A3'] = 'Risk Level'
    ws1['B3'] = data.get('risk_level', '')
    ws1['A4'] = 'Risk Probability'
    ws1['B4'] = data.get('risk_percentage', '')
    ws1['A5'] = 'Suggestion'
    ws1['B5'] = data.get('suggestion', '')

    ws2 = wb.create_sheet('Input Data')
    ws2['A1'] = 'Parameter'
    ws2['B1'] = 'Value'

    field_map = {
        'age': 'Age',
        'gender': 'Gender',
        'daily_social_media_hours': 'Daily Social Media (hours)',
        'platform_usage': 'Platform',
        'screen_time_before_sleep': 'Screen Time Before Sleep (hours)',
        'sleep_hours': 'Sleep (hours)',
        'physical_activity': 'Physical Activity',
        'social_interaction_level': 'Social Interaction',
        'academic_performance': 'Academic Performance',
        'stress_level': 'Stress Level',
        'anxiety_level': 'Anxiety Level',
        'addiction_level': 'Addiction Level'
    }

    gender_map = {'male': 'Male', 'female': 'Female'}
    activity_map = {0: 'None', 1: 'Occasional', 2: 'Regular'}
    social_map = {'low': 'Low', 'medium': 'Medium', 'high': 'High'}
    academic_map = {1: 'Poor', 2: 'Average', 3: 'Good', 4: 'Excellent'}

    row = 2
    for key, value in data.get('input_data', {}).items():
        label = field_map.get(key, key)
        if key == 'gender':
            value = gender_map.get(value, value)
        elif key == 'physical_activity':
            value = activity_map.get(value, value)
        elif key == 'social_interaction_level':
            value = social_map.get(value, value)
        elif key == 'academic_performance':
            value = academic_map.get(value, value)
        ws2[f'A{row}'] = label
        ws2[f'B{row}'] = value
        row += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


def export_to_pdf(data, filename='prediction_report.pdf'):
    """导出到PDF（纯英文版，无中文）"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName='Helvetica',
        fontSize=20,
        textColor=colors.HexColor('#FF6600'),
        alignment=1,
        spaceAfter=20
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontName='Helvetica',
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceBefore=12,
        spaceAfter=6
    )

    story = []

    story.append(Paragraph('Depression Risk Prediction Report', title_style))
    story.append(Spacer(1, 12))

    story.append(Paragraph('1. Prediction Result', heading_style))
    story.append(Spacer(1, 6))

    # 处理建议（英文）
    suggestion = data.get('suggestion', '')
    if '高风险' in suggestion:
        suggestion = 'High Risk: Seek professional psychological counseling as soon as possible.'
    elif '中风险' in suggestion:
        suggestion = 'Medium Risk: Pay attention to self-regulation and maintain regular routine.'
    else:
        suggestion = 'Low Risk: Maintain healthy living habits.'

    result_data = [
        ['Risk Level', data.get('risk_level', '')],
        ['Risk Probability', data.get('risk_percentage', '')],
        ['Suggestion', suggestion]
    ]

    result_table = Table(result_data, colWidths=[50*mm, 100*mm])
    result_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(result_table)
    story.append(Spacer(1, 12))

    story.append(Paragraph('2. Input Data', heading_style))
    story.append(Spacer(1, 6))

    field_map = {
        'age': 'Age',
        'gender': 'Gender',
        'daily_social_media_hours': 'Daily Social Media (hours)',
        'platform_usage': 'Platform',
        'screen_time_before_sleep': 'Screen Time Before Sleep (hours)',
        'sleep_hours': 'Sleep (hours)',
        'physical_activity': 'Physical Activity',
        'social_interaction_level': 'Social Interaction',
        'academic_performance': 'Academic Performance',
        'stress_level': 'Stress Level',
        'anxiety_level': 'Anxiety Level',
        'addiction_level': 'Addiction Level'
    }

    gender_map = {'male': 'Male', 'female': 'Female'}
    activity_map = {0: 'None', 1: 'Occasional', 2: 'Regular'}
    social_map = {'low': 'Low', 'medium': 'Medium', 'high': 'High'}
    academic_map = {1: 'Poor', 2: 'Average', 3: 'Good', 4: 'Excellent'}

    input_data = []
    for key, value in data.get('input_data', {}).items():
        label = field_map.get(key, key)
        if key == 'gender':
            value = gender_map.get(value, value)
        elif key == 'physical_activity':
            value = activity_map.get(value, value)
        elif key == 'social_interaction_level':
            value = social_map.get(value, value)
        elif key == 'academic_performance':
            value = academic_map.get(value, value)
        input_data.append([label, str(value)])

    input_table = Table(input_data, colWidths=[60*mm, 90*mm])
    input_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(input_table)

    doc.build(story)
    return response